import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app, session
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from typing_extensions import Annotated
from src.models import db, User, Transaction, InitialBalance, UserPreferences
from src.forms import LoginForm, RegistrationForm
from src.anthropic_service import FinancialAnalytics
from src.upload_handler import process_upload
from src.utils import calculate_totals, calculate_burn_rate, calculate_runway
from src.config import Config
import pandas as pd
from io import BytesIO
from dotenv import load_dotenv
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import io
import base64
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from matplotlib.dates import MonthLocator, DateFormatter
import matplotlib.ticker as ticker
from datetime import datetime
import calendar
from flask_migrate import Migrate
from flask_babel import Babel

# Load .env file explicitly at the start
load_dotenv()

# Ensure instance folder exists
instance_path = os.path.join(os.path.dirname(__file__),'instance')
os.makedirs(instance_path, exist_ok=True)

upload_path = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(upload_path, exist_ok=True)

app = Flask(__name__,
            template_folder='templates',
            static_folder='static')
app.config.from_object(Config)

db.init_app(app)

migrate = Migrate(app, db)

with app.app_context():
    db.create_all()

login_manager = LoginManager(app)
login_manager.login_view = 'login'

babel = Babel()

def get_locale():
    # Debug print
    print(f"get_locale called, session: {session}")
    
    # First check if a language is stored in the session
    if 'language' in session:
        lang = session['language']
        print(f"Using language from session: {lang}")
        return lang
        
    # Otherwise fallback to browser preference
    browser_lang = request.accept_languages.best_match(['en', 'es', 'ja'])
    print(f"Using browser language: {browser_lang}")
    return browser_lang

# Initialize babel with the locale selector function
babel.init_app(app, locale_selector=get_locale)

@app.route('/set-language/<language>')
def set_language(language):
    # Store language in session
    session['language'] = language
    print(f"Setting language to: {language}")
    print(f"Session contains: {session}")
    # Debug response to confirm language setting
    flash(f'Language set to: {language}', 'info')
    # Redirect back to the referring page or home page
    return redirect(request.referrer or url_for('home'))

@login_manager.user_loader
def load_user(user_id):
    #return User.query.get(int(user_id))
    return db.session.get(User, int(user_id))

#All routes in app
@app.route('/',methods=['GET','POST'])
@app.route('/home', methods=['GET', 'POST'])
@login_required
def home():
    user_status = "Logged In"
    page = request.args.get('page', 1, type=int)
    per_page = 8

    if request.method == 'POST':
        # Handle adding transactions
        new_transaction = Transaction(
            user_id=current_user.id,
            date=request.form.get('date'),
            description=request.form.get('description'),
            amount=float(request.form.get('amount')),
            type=request.form.get('type')
        )
        db.session.add(new_transaction)
        db.session.commit()
        flash('Transaction added successfully', 'success')
        return redirect(url_for('home'))

    # Filter transactions by current user
    paginated_transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date.desc()).paginate(page=page, per_page=per_page)
    all_transactions = Transaction.query.filter_by(user_id=current_user.id).all()
    
    # Get initial balance for current user
    initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
    initial_balance = initial_balance_record.balance if initial_balance_record else 0.0

    total_cfo, total_cfi, total_cff = calculate_totals(all_transactions)
    balance = initial_balance + total_cfo + total_cfi + total_cff

    return render_template('home.html', transactions=paginated_transactions, balance=balance, initial_balance=initial_balance,
                           total_cfo=total_cfo, total_cfi=total_cfi, total_cff=total_cff, user=current_user, status=user_status)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and check_password_hash(user.password, form.password.data):
            login_user(user, remember=True)
            return redirect(url_for('home'))
        else:
            flash('Login Unsuccessful. Please check username and password', 'danger')
    return render_template('login.html', form=form)

@app.route('/register',methods=['GET','POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = RegistrationForm()
    if form.validate_on_submit():
        user_exists = User.query.filter_by(username=form.username.data).first()
        if user_exists:
            flash('Username already exists. Please choose a different one.', 'danger')
        else:
            hashed_password = generate_password_hash(form.password.data)
            new_user = User(username=form.username.data, password=hashed_password)
            db.session.add(new_user)
            db.session.commit()
            new_preferences = UserPreferences(user_id=new_user.id)
            db.session.add(new_preferences)
            db.session.commit()
            flash('Your account has been created! You can now log in.', 'success')
            return redirect(url_for('login'))
    return render_template('register.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/edit/<int:transaction_id>', methods=['GET', 'POST'])
@login_required
def edit_transaction(transaction_id):
    transaction = Transaction.query.get_or_404(transaction_id)
    if request.method == 'POST':
        transaction.date = request.form['date']
        transaction.description = request.form['description']
        transaction.amount = float(request.form['amount'])
        transaction.type = request.form['type']
        db.session.commit()
        flash('Transaction Updated Successfully', 'success')
        return redirect(url_for('cash_activities'))
    return render_template('edit_transaction.html', transaction=transaction)

@app.route('/delete/<int:transaction_id>', methods=['POST'])
@login_required
def delete_transaction(transaction_id):
    transaction = Transaction.query.get_or_404(transaction_id)
    db.session.delete(transaction)
    db.session.commit()
    flash('Your Transaction Deleted!', 'danger')
    return redirect(url_for('home'))

@app.route('/upload', methods=['GET', 'POST'])
@login_required
def upload_route():
    if request.method == 'GET':
        return render_template('upload.html')
    elif request.method == 'POST':
        if 'file' not in request.files:
            return jsonify({'error': 'No file part'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No selected file'}), 400
        
        try:
            result = process_upload(file)
            return jsonify(result)
        except Exception as e:
            app.logger.error(f"Error in upload_file: {str(e)}")
            return jsonify({'error': str(e)}), 500

@app.route('/set-initial-balance', methods=['POST'])
@login_required
def set_initial_balance():
    try:
        amount = float(request.form.get('initial_balance'))
        
        # Get or create initial balance record for the user
        initial_balance = InitialBalance.query.filter_by(user_id=current_user.id).first()
        if initial_balance:
            initial_balance.balance = amount
        else:
            initial_balance = InitialBalance(user_id=current_user.id, balance=amount)
            db.session.add(initial_balance)
            
        db.session.commit()
        flash('Initial balance set successfully', 'success')
    except ValueError:
        flash('Please enter a valid number for the initial balance', 'danger')
    except Exception as e:
        flash(f'Error setting initial balance: {str(e)}', 'danger')
        
    return redirect(url_for('cash_overview'))

@app.route('/export/<file_type>')
@login_required
def export(file_type):
    transactions = Transaction.query.all()
    data = [{
        'Date': transaction.date,
        'Description': transaction.description,
        'Amount': transaction.amount,
        'Type': transaction.type
    } for transaction in transactions]

    df = pd.DataFrame(data)

    if file_type == 'csv':
        output = BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        return send_file(output, mimetype='text/csv', as_attachment=True, download_name='transactions.csv')
    elif file_type == 'excel':
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Transactions')
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                         as_attachment=True, download_name='transactions.xlsx')
    else:
        flash('Invalid file type requested.', 'danger')
        return redirect(url_for('home'))

@app.route('/save_transactions', methods=['POST'])
@login_required
def save_transactions():
    data = request.json
    try:
        for item in data:
            new_transaction = Transaction(
                user_id=current_user.id,
                date=item['date'],
                description=item['description'],
                amount=float(item['amount']),
                type=item['type']
            )
            db.session.add(new_transaction)
        db.session.commit()
        return jsonify({'message': 'Transactions saved successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/balance-by-date', methods=['POST'])
@login_required
def balance_by_date():
    try:
        date_str = request.form.get('date')
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Get initial balance
        initial_balance = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_amount = initial_balance.balance if initial_balance else 0
        
        # Get all transactions up to the target date
        transactions = Transaction.query.filter(
            Transaction.user_id == current_user.id,
            Transaction.date <= target_date
        ).all()
        
        # Calculate totals
        total_cfo, total_cfi, total_cff = calculate_totals(transactions)
        balance = initial_amount + total_cfo + total_cfi + total_cff
        
        return jsonify({
            'success': True,
            'balance': balance,
            'date': date_str
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    
@app.route('/forecast', methods=['GET'])
@login_required
def forecast():
    try:
        # Get all transactions for the current user
        transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
        
        if not transactions:
            return jsonify({
                'error': 'No transactions found. Please add some transactions first.'
            }), 400
            
        # Convert transactions to a format suitable for analysis
        transaction_data = [{
            'date': t.date.strftime('%Y-%m-%d') if isinstance(t.date, datetime) else t.date,
            'amount': t.amount,
            'type': t.type,
            'description': t.description
        } for t in transactions]
        
        # Initialize financial analytics with API key
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({
                'error': 'API key not found. Please set the ANTHROPIC_API_KEY environment variable.'
            }), 500
            
        analytics = FinancialAnalytics(api_key=api_key)
        
        # Get initial balance
        initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_balance = initial_balance_record.balance if initial_balance_record else 0.0
        
        # Calculate current balance
        current_balance = initial_balance + sum(t.amount for t in transactions)
        
        # Mock working capital data (you may want to replace this with actual data)
        working_capital = {
            'current_assets': current_balance,
            'current_liabilities': 0,
            'cash': current_balance
        }
        
        # Get analysis results using the correct method
        analysis_results = analytics.generate_advanced_financial_analysis(
            initial_balance=initial_balance,
            current_balance=current_balance,
            transaction_history=transaction_data,
            working_capital=working_capital
        )
        
        return jsonify({
            'success': True,
            'ai_analysis': analysis_results.get('ai_analysis', 'No insights available'),
            'patterns': analysis_results.get('patterns', {'seasonal_pattern': [0] * 12}),
            'forecasts': analysis_results.get('forecasts', {'90_days': [0] * 90}),
            'risk_metrics': analysis_results.get('risk_metrics', {
                'liquidity_ratio': 0,
                'cash_flow_volatility': 0,
                'burn_rate': 0,
                'runway_months': 0
            })
        })
    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500

@app.route('/generate_cashflow_statement', methods=['GET'])
@login_required
def generate_cashflow_statement_route():
    try:
        transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
        if not transactions:
            app.logger.warning("No transactions found when generating cash flow statement")
            return jsonify({'error': 'No transactions found'}), 400

        initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_balance = initial_balance_record.balance if initial_balance_record else 0.0

        start_date = transactions[0].date
        end_date = transactions[-1].date

        transaction_data = "\n".join([f"Date: {t.date}, Description: {t.description}, Amount: {t.amount}, Type: {t.type}" for t in transactions])

        app.logger.info(f"Generating cash flow statement for period {start_date} to {end_date}")
        
        # Initialize FinancialAnalytics
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            app.logger.error("ANTHROPIC_API_KEY not found in environment variables")
            return jsonify({'error': 'API key not found'}), 500
            
        analytics = FinancialAnalytics(api_key=api_key)
        
        # Generate cash flow statement using AI
        statement_data, ending_balance = analytics.generate_cashflow_statement(
            initial_balance=initial_balance,
            start_date=start_date,
            end_date=end_date,
            transaction_data=transaction_data
        )
        
        if not statement_data:
            app.logger.error("generate_cashflow_statement returned None")
            return jsonify({'error': 'Failed to generate statement: No data returned'}), 500

        app.logger.info("Cash flow statement generated successfully")
        app.logger.debug(f"Statement data: {statement_data}")

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow Statement"

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15

        # Add title
        ws['A1'] = f"Cash Flow Statement - {start_date} to {end_date}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        # Add headers
        headers = ['Category', 'Amount']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Helper function to add a row with formatting
        def add_row(row, category, amount, is_total=False, indent=0):
            ws.cell(row=row, column=1, value=category).alignment = Alignment(indent=indent)
            ws.cell(row=row, column=2, value=amount)
            if is_total:
                for col in range(1, 3):
                    ws.cell(row=row, column=col).font = Font(bold=True)

        # Add Beginning Cash Balance
        row = 4
        add_row(row, "Beginning Cash Balance", initial_balance, True)
        row += 2

        # Process and add data for each category
        categories = ['CFO', 'CFI', 'CFF']
        for category in categories:
            ws.cell(row=row, column=1, value=f"Cash Flow from {'Operating' if category == 'CFO' else 'Investing' if category == 'CFI' else 'Financing'} Activities ({category})")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            row += 1

            category_total = 0
            for item in statement_data:
                if item['Category'] == category:
                    add_row(row, item['Subcategory'], item['Amount'], indent=1)
                    category_total += item['Amount']
                    row += 1

            add_row(row, f"Net Cash from {'Operating' if category == 'CFO' else 'Investing' if category == 'CFI' else 'Financing'}", category_total, True)
            row += 2

        # Add Total Net Cash Flow
        total_net_cash_flow = sum(item['Amount'] for item in statement_data)
        add_row(row, "Total Net Cash Flow", total_net_cash_flow, True)
        row += 2

        # Add Ending Cash Balance
        ending_balance = initial_balance + total_net_cash_flow
        add_row(row, "Ending Cash Balance", ending_balance, True)

        # Apply currency formatting to amount column
        for row in ws['B4:B' + str(ws.max_row)]:
            for cell in row:
                cell.number_format = '#,##0.00'

        # Apply right alignment to amount column
        for row in ws['B1:B' + str(ws.max_row)]:
            for cell in row:
                cell.alignment = Alignment(horizontal='right')

        # Add borders
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

        # Save to BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        app.logger.info("Excel file created successfully")
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='cash_flow_statement.xlsx'
        )
    except Exception as e:
        app.logger.error(f"Failed to generate cash flow statement: {str(e)}")
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Failed to generate cash flow statement: {str(e)}'}), 500

# Generate monthly chart function    
def generate_monthly_balance_chart():
    # Get transactions for the current user
    transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
    
    # Get initial balance for the current user
    initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
    initial_balance = initial_balance_record.balance if initial_balance_record else 0.0

    monthly_balances = []
    current_balance = initial_balance
    current_month = None

    for transaction in transactions:
        transaction_date = transaction.date if isinstance(transaction.date, datetime) else datetime.strptime(transaction.date, '%Y-%m-%d')
        
        if current_month != transaction_date.replace(day=1):
            if current_month:
                last_day = calendar.monthrange(current_month.year, current_month.month)[1]
                monthly_balances.append({
                    'date': current_month.replace(day=last_day),
                    'balance': current_balance
                })
            current_month = transaction_date.replace(day=1)

        current_balance += transaction.amount

    # Add the last month's balance
    if current_month:
        last_day = calendar.monthrange(current_month.year, current_month.month)[1]
        monthly_balances.append({
            'date': current_month.replace(day=last_day),
            'balance': current_balance
        })

    # Create the chart
    fig = Figure(figsize=(12, 6))
    axis = fig.add_subplot(1, 1, 1)
    
    if monthly_balances:
        dates = [balance['date'] for balance in monthly_balances]
        balances = [balance['balance'] for balance in monthly_balances]
        axis.plot(dates, balances, marker='o')  # Added markers for each data point

        # Improve Y-axis formatting
        def currency_formatter(x, p):
            return f'${x:,.0f}'
        
        axis.yaxis.set_major_formatter(ticker.FuncFormatter(currency_formatter))
        
        # Adjust Y-axis ticks for better readability
        axis.yaxis.set_major_locator(ticker.MaxNLocator(nbins=10, integer=True))
        
        # Format X-axis to show dates nicely
        axis.xaxis.set_major_locator(MonthLocator())
        axis.xaxis.set_major_formatter(DateFormatter('%Y-%m-%d'))
        
        # Add some padding to the y-axis
        ylim = axis.get_ylim()
        axis.set_ylim([ylim[0] - (ylim[1] - ylim[0]) * 0.1, ylim[1] + (ylim[1] - ylim[0]) * 0.1])
    else:
        # If no data, show a message on the chart
        axis.text(0.5, 0.5, 'No transaction data available', 
                 horizontalalignment='center', verticalalignment='center',
                 transform=axis.transAxes, fontsize=14)
        axis.set_xticks([])
        axis.set_yticks([])

    axis.set_title('Monthly Balance', fontsize=16, fontweight='bold')
    axis.set_xlabel('Date', fontsize=12)
    axis.set_ylabel('Balance ($)', fontsize=12)
    axis.tick_params(axis='both', which='major', labelsize=10)
    axis.tick_params(axis='x', rotation=45)
    
    # Add gridlines for better readability
    axis.grid(True, linestyle='--', alpha=0.7)

    fig.tight_layout()

    # Convert plot to PNG image
    png_image = io.BytesIO()
    FigureCanvas(fig).print_png(png_image)
    
    # Encode PNG image to base64 string
    png_image_b64_string = "data:image/png;base64,"
    png_image_b64_string += base64.b64encode(png_image.getvalue()).decode('utf8')

    return png_image_b64_string

#Map plotting chart
@app.route('/monthly-balances', methods=['GET'])
@login_required
def monthly_balances():
    try:
        app.logger.info("Starting to generate monthly balance data")
        transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
        initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_balance = initial_balance_record.balance if initial_balance_record else 0.0

        monthly_data = {}
        current_balance = initial_balance

        for transaction in transactions:
            month = transaction.date.strftime('%Y-%m') if isinstance(transaction.date, datetime) else datetime.strptime(transaction.date, '%Y-%m-%d').strftime('%Y-%m')
            
            if month not in monthly_data:
                monthly_data[month] = current_balance
            
            current_balance += transaction.amount
            monthly_data[month] = current_balance

        app.logger.info(f"Monthly balance data generated: {monthly_data}")
        return jsonify({
            'success': True,
            'data': {
                'labels': list(monthly_data.keys()),
                'balances': list(monthly_data.values())
            }
        })
    except Exception as e:
        app.logger.error(f"Error generating monthly balance data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/settings')
@login_required
def settings():
    user_preferences = UserPreferences.query.filter_by(user_id=current_user.id).first()
    return render_template('settings.html', user_preferences=user_preferences)

@app.route('/update_profile', methods=['POST'])
@login_required
def update_profile():
    current_user.username = request.form.get('username')
    current_user.email = request.form.get('email')
    db.session.commit()
    flash('Profile updated successfully', 'success')
    return redirect(url_for('settings'))

@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    if check_password_hash(current_user.password, request.form.get('current_password')):
        if request.form.get('new_password') == request.form.get('confirm_password'):
            current_user.password = generate_password_hash(request.form.get('new_password'))
            db.session.commit()
            flash('Password changed successfully', 'success')
        else:
            flash('New passwords do not match', 'danger')
    else:
        flash('Current password is incorrect', 'danger')
    return redirect(url_for('settings'))

@app.route('/update_modules', methods=['POST'])
@login_required
def update_modules():
    user_preferences = UserPreferences.query.filter_by(user_id=current_user.id).first()
    if not user_preferences:
        user_preferences = UserPreferences(user_id=current_user.id)
        db.session.add(user_preferences)
    user_preferences.modules = request.form.getlist('modules')
    db.session.commit()
    flash('Module preferences updated successfully', 'success')
    return redirect(url_for('settings'))

@app.context_processor
def utility_processor():
    def get_user_preferences():
        if current_user.is_authenticated:
            prefs = UserPreferences.query.filter_by(user_id=current_user.id).first()
            if not prefs:
                prefs = UserPreferences(user_id=current_user.id)
                db.session.add(prefs)
                db.session.commit()
            return prefs
        return None
    return dict(user_preferences=get_user_preferences())

@app.route('/cash-overview')
@login_required
def cash_overview():
    # Get or create initial balance for the current user
    initial_balance = InitialBalance.query.filter_by(user_id=current_user.id).first()
    if not initial_balance:
        initial_balance = InitialBalance(user_id=current_user.id, balance=0)
        db.session.add(initial_balance)
        db.session.commit()

    # Get all transactions for the current user
    transactions = Transaction.query.filter_by(user_id=current_user.id).all()
    
    # Calculate totals
    total_cfo, total_cfi, total_cff = calculate_totals(transactions)
    balance = initial_balance.balance + total_cfo + total_cfi + total_cff

    # Calculate burn rate and runway
    burn_rate = calculate_burn_rate(transactions)
    runway_months = calculate_runway(balance, burn_rate)

    return render_template('cash_overview.html', 
                         initial_balance=initial_balance.balance,
                         total_cfo=total_cfo, 
                         total_cfi=total_cfi, 
                         total_cff=total_cff, 
                         balance=balance,
                         burn_rate=burn_rate,
                         runway_months=runway_months)

@app.route('/cash-activities')
@login_required
def cash_activities():
    page = request.args.get('page', 1, type=int)
    per_page = 8
    
    # Get paginated transactions for the current user
    paginated_transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date.desc()).paginate(page=page, per_page=per_page)
    
    return render_template('cash_activities.html', transactions=paginated_transactions)

@app.route('/ai-analysis')
@login_required
def ai_analysis():
    return render_template('ai_analysis.html')

@app.route('/create-transaction', methods=['POST'])
@login_required
def create_transaction():
    try:
        new_transaction = Transaction(
            user_id=current_user.id,
            date=request.form.get('date'),
            description=request.form.get('description'),
            amount=float(request.form.get('amount')),
            type=request.form.get('type')
        )
        db.session.add(new_transaction)
        db.session.commit()
        flash('Transaction added successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error adding transaction: ' + str(e), 'danger')
    
    return redirect(url_for('cash_activities'))

@app.route('/monthly-income-expense', methods=['GET'])
@login_required
def monthly_income_expense():
    try:
        transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
        monthly_data = {}

        for transaction in transactions:
            month = transaction.date.strftime('%Y-%m') if isinstance(transaction.date, datetime) else datetime.strptime(transaction.date, '%Y-%m-%d').strftime('%Y-%m')
            
            if month not in monthly_data:
                monthly_data[month] = {'income': 0, 'expense': 0}
            
            if transaction.amount > 0:
                monthly_data[month]['income'] += transaction.amount
            else:
                monthly_data[month]['expense'] += abs(transaction.amount)

        return jsonify({
            'success': True,
            'data': {
                'labels': list(monthly_data.keys()),
                'income': [data['income'] for data in monthly_data.values()],
                'expense': [data['expense'] for data in monthly_data.values()]
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/cashout-categories', methods=['GET'])
@login_required
def cashout_categories():
    try:
        # Fix the filter syntax for negative amounts
        transactions = Transaction.query.filter(
            Transaction.user_id == current_user.id,
            Transaction.amount < 0  # Correct syntax for filtering negative amounts
        ).all()

        categories_data = {}
        for transaction in transactions:
            category = transaction.type
            if category not in categories_data:
                categories_data[category] = 0
            categories_data[category] += abs(transaction.amount)

        return jsonify({
            'success': True,
            'data': {
                'labels': list(categories_data.keys()),
                'amounts': list(categories_data.values())
            }
        })
    except Exception as e:
        app.logger.error(f"Error fetching cashout categories: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/debug-language')
def debug_language():
    """Debug endpoint to check translations"""
    all_info = {
        'session_language': session.get('language', 'Not set'),
        'available_translations': os.listdir('translations'),
        'session_data': dict(session),
        'best_match': request.accept_languages.best_match(['en', 'es', 'ja']),
    }
    return jsonify(all_info)

if __name__ == '__main__':
    app.run(debug=True)